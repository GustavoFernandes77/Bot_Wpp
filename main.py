import tkinter as tk
from tkinter import ttk
import openpyxl
import os
from datetime import datetime

# Caminho do novo arquivo Excel
path = r"'''Caminho do arquivo Excel'''"
# Comentario Edgar
def create_new_file():
    """Cria um novo arquivo Excel com os cabeçalhos se o arquivo não existir"""
    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Definindo os cabeçalhos
        sheet.append(["Nome", "Telefone", "Cpf", "Email", "Endereco", "Vencimento", "Pagou"])
        workbook.save(path)

def load_data():
    """Carrega dados do arquivo Excel ou cria um novo arquivo se necessário"""
    create_new_file()
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Obtem os valores da planilha
    list_values = list(sheet.values)

    # Filtra cabeçalhos preenchidos na primeira linha
    column_names = [str(col) for col in list_values[0] if col is not None]

    # Configura dinamicamente as colunas no Treeview
    treeview["columns"] = column_names

    # Define cabeçalhos e colunas
    for col_name in column_names:
        treeview.heading(col_name, text=col_name)  # Define cabeçalho visível
        treeview.column(col_name, width=100)  # Configura largura padrão

    # Adiciona as linhas preenchidas
    for value_tuple in list_values[1:]:
        # Ignora linhas onde todos os valores são None ou vazios
        if any(value is not None and str(value).strip() != "" for value in value_tuple):
            treeview.insert('', tk.END, values=value_tuple[:len(column_names)])  # Insere apenas os valores correspondentes às colunas

def insert_row():
    name = name_entry.get()
    telefone = telefone_entry.get()
    cpf = cpf_entry.get()
    email = email_entry.get()
    endereco = endereco_entry.get()
    vencimento = vencimento_entry.get()
    pagou = "Pagou" if a.get() else "Não Pagou"

    print("Nome:", name, "Cpf:", cpf, "Telefone:", telefone, "Email:", email, "Endereço:", endereco, "Data de Vencimento:", vencimento, "Pagou:", pagou)

    # Insert row into Excel sheet
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [name, telefone, cpf, email, endereco, vencimento, pagou]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)
    
    # Clear the values
    name_entry.delete(0, "end")
    name_entry.insert(0, "Nome")

    telefone_entry.delete(0, "end")
    telefone_entry.insert(0, "Telefone")

    cpf_entry.delete(0, "end")
    cpf_entry.insert(0, "Cpf")

    email_entry.delete(0, "end")
    email_entry.insert(0, "Email")

    endereco_entry.delete(0, "end")
    endereco_entry.insert(0, "Endereço")

    vencimento_entry.delete(0, "end")
    vencimento_entry.insert(0, "Vencimento")

    checkbutton.state(["!selected"])

def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

def remove_row(tree):
    selected_item = tree.selection()  # Obtém o item selecionado
    if not selected_item:
        return  # Nada selecionado, não faz nada

    # Obtém o índice da linha selecionada
    selected_index = tree.index(selected_item)

    # Remove a linha do Treeview
    tree.delete(selected_item)

    # Remove a linha do arquivo Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Remove a linha correspondente no Excel (índice + 2, pois a primeira linha é o cabeçalho)
    sheet.delete_rows(selected_index + 2)
    workbook.save(path)

def edit_row(tree):
    selected_item = tree.selection()  # Obtém o item selecionado
    if not selected_item:
        return  # Nada selecionado, não faz nada

    # Obtém os valores da linha selecionada
    item_values = tree.item(selected_item, 'values')

    # Preenche os campos de entrada com os valores da linha selecionada
    name_entry.delete(0, "end")
    name_entry.insert(0, item_values[0])

    telefone_entry.delete(0, "end")
    telefone_entry.insert(0, item_values[1])

    cpf_entry.delete(0, "end")
    cpf_entry.insert(0, item_values[2])

    email_entry.delete(0, "end")
    email_entry.insert(0, item_values[3])

    endereco_entry.delete(0, "end")
    endereco_entry.insert(0, item_values[4])

    vencimento_entry.delete(0, "end")
    vencimento_entry.insert(0, item_values[5])

    a.set(True if item_values[6] == "Pagou" else False)

    # Altera o botão "Inserir" para "Salvar Edição"
    button.config(text="Salvar Edição", command=lambda: save_edit(tree, selected_item))

def save_edit(tree, selected_item):
    # Obtém os novos valores dos campos de entrada
    name = name_entry.get()
    telefone = telefone_entry.get()
    cpf = cpf_entry.get()
    email = email_entry.get()
    endereco = endereco_entry.get()
    vencimento = vencimento_entry.get()
    pagou = "Pagou" if a.get() else "Não Pagou"

    # Atualiza a linha no Treeview
    tree.item(selected_item, values=(name, telefone, cpf, email, endereco, vencimento, pagou))

    # Atualiza a linha no arquivo Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Obtém o índice da linha selecionada
    selected_index = tree.index(selected_item)

    # Atualiza a linha correspondente no Excel (índice + 2, pois a primeira linha é o cabeçalho)
    row = selected_index + 2
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=telefone)
    sheet.cell(row=row, column=3, value=cpf)
    sheet.cell(row=row, column=4, value=email)
    sheet.cell(row=row, column=5, value=endereco)
    sheet.cell(row=row, column=6, value=vencimento)
    sheet.cell(row=row, column=7, value=pagou)

    workbook.save(path)

    # Restaura o botão "Inserir" ao estado original
    button.config(text="Inserir", command=insert_row)

    # Limpa os campos de entrada
    name_entry.delete(0, "end")
    name_entry.insert(0, "Nome")

    telefone_entry.delete(0, "end")
    telefone_entry.insert(0, "Telefone")

    cpf_entry.delete(0, "end")
    cpf_entry.insert(0, "Cpf")

    email_entry.delete(0, "end")
    email_entry.insert(0, "Email")

    endereco_entry.delete(0, "end")
    endereco_entry.insert(0, "Endereço")

    vencimento_entry.delete(0, "end")
    vencimento_entry.insert(0, "Vencimento")

    checkbutton.state(["!selected"])

def clear_placeholder(event, entry, placeholder):
    """Limpa o campo de entrada apenas se ele contiver o texto de placeholder"""
    if entry.get() == placeholder:
        entry.delete(0, "end")

root = tk.Tk()

style = ttk.Style(root)
theme_path = os.path.join(os.path.dirname(__file__), "forest-light.tcl")
root.tk.call("source", theme_path)
theme_path2 = os.path.join(os.path.dirname(__file__), "forest-dark.tcl")
root.tk.call("source", theme_path2)
style.theme_use("forest-dark")

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Inserir Dados")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Nome")
name_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, name_entry, "Nome"))
name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

telefone_entry = ttk.Entry(widgets_frame)
telefone_entry.insert(0, "Telefone")
telefone_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, telefone_entry, "Telefone"))
telefone_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

cpf_entry = ttk.Entry(widgets_frame)
cpf_entry.insert(0, "Cpf")
cpf_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, cpf_entry, "Cpf"))
cpf_entry.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

email_entry = ttk.Entry(widgets_frame)
email_entry.insert(0, "Email")
email_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, email_entry, "Email"))
email_entry.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="ew")

endereco_entry = ttk.Entry(widgets_frame)
endereco_entry.insert(0, "Endereço")
endereco_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, endereco_entry, "Endereço"))
endereco_entry.grid(row=4, column=0, padx=5, pady=(0, 5), sticky="ew")

vencimento_entry = ttk.Entry(widgets_frame)
vencimento_entry.insert(0, "Vencimento")
vencimento_entry.bind("<FocusIn>", lambda e: clear_placeholder(e, vencimento_entry, "Vencimento"))
vencimento_entry.grid(row=5, column=0, padx=5, pady=(0, 5), sticky="ew")

a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widgets_frame, text="Pagou", variable=a)
checkbutton.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

button = ttk.Button(widgets_frame, text="Inserir", command=insert_row)
button.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")

separator = ttk.Separator(widgets_frame)
separator.grid(row=8, column=0, padx=(20, 10), pady=10, sticky="ew")

mode_switch = ttk.Checkbutton(
    widgets_frame, text="Modo", style="Switch", command=toggle_mode)
mode_switch.grid(row=9, column=0, padx=5, pady=10, sticky="nsew")

# Adicionar botões de editar e remover
edit_button = ttk.Button(widgets_frame, text="Editar", command=lambda: edit_row(treeview))
edit_button.grid(row=10, column=0, padx=5, pady=5, sticky="nsew")

remove_button = ttk.Button(widgets_frame, text="Remover", command=lambda: remove_row(treeview))
remove_button.grid(row=11, column=0, padx=5, pady=5, sticky="nsew")

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Nome", "Telefone", "Cpf", "Email", "Endereco", "Vencimento", "Pagou")
treeview = ttk.Treeview(treeFrame, show="headings",
yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Nome", width=100)
treeview.column("Telefone", width=50)
treeview.column("Cpf", width=100)
treeview.column("Email", width=100)
treeview.column("Endereco", width=100)
treeview.column("Vencimento", width=100)
treeview.column("Pagou", width=50)
treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

root.mainloop()
